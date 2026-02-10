import numpy as np
import math
import pandas as pd
import openpyxl
import streamlit as st
import pandas as pd
import altair as alt
import base64
import io
from PIL import Image
import json
import os
import glob
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.message import EmailMessage
icon = Image.open("betekoppen_logo.png")
st.set_page_config(page_title="Jureren Betekoppen",
                  page_icon=icon,
                  layout="wide")
#%%
# INITIALISATIE VAN STATE
def init_session():
    """Zorgt dat session_state altijd consistente waarden heeft."""
    defaults = {
        "logged_in": False,
        "username": None,
        "soort": None,
        "active_tab": "Home",
        "pending_saves": [],
        
        "df_beoordelingen_cache": None, # alg. beoordelingen
        "sheet_beoordelingen": None, # alg. beoordelingen
        
        "df_top3_cache": None, # leutigste deelnemer
        "sheet_top3": None,
        
        "uitslag_berekend" : False,
        "df_rapport" : None,
        "df_pers" : None,
        "Rapport_excel": None,
        "Pers_excel":None,
        "mail_verzonden": False
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value
# Functie direct aanroepen
init_session()

#%% GOOGLES SHEETS SETUP

scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds_dict = st.secrets["gcp_service_account"]
creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
client = gspread.authorize(creds)

## Open de Google Sheet
# sheet = client.open("Jury_beoordelingen_2026_v1").worksheet("Beoordeling")
# sheet_top3 = client.open("Jury_beoordelingen_2026_v1").worksheet("LeutigsteDeelnemer")

# Bestaande data ophalen
@st.cache_data(ttl=300)
def load_sheet_data(sheet_name):
    sheet = client.open("Jury_beoordelingen_2026_v1").worksheet(sheet_name)
    records = sheet.get_all_records()
    df = pd.DataFrame(records)
    df.columns = [c.strip() for c in df.columns]
    return df

# google sheets 1x per sessie openen
if st.session_state.sheet_beoordelingen is None:
    st.session_state.sheet_beoordelingen = (client.open("Jury_beoordelingen_2026_v1").worksheet("Beoordelingen_2026"))
if st.session_state.df_beoordelingen_cache is None:
    st.session_state.df_beoordelingen_cache = load_sheet_data("Beoordelingen_2026")
    
if st.session_state.sheet_top3 is None:
    st.session_state.sheet_top3 = (client.open("Jury_beoordelingen_2026_v1").worksheet("LeutigsteDeelnemer_2026"))
if st.session_state.df_top3_cache is None:
    st.session_state.df_top3_cache = load_sheet_data("LeutigsteDeelnemer_2026")
#%%
def mail_excel(excel_bytes_1, filename_1, excel_bytes_2, filename_2):
    msg = EmailMessage()
    msg["Subject"] = "Uitslag carnavalsoptocht Sas van Gent (Betekoppen) 2026 [geautomatiseerde mail]"
    msg["From"] = st.secrets["email"]["from"]
    msg["To"] = st.secrets["email"]["to"]
    msg["Cc"] = "kasper.tak12@gmail.com" 
    
    msg.set_content(
        "Beste, \n\nAlle onderdelen zijn beoordeeld door de juryleden.\n"
        "\nIn de bijlage zijn twee Excelbestanden te vinden. Het rapport bevat de totale, gedetailleerde beoordeling. Het andere bestand is geschikt voor de pers."
        "\nVia deze link kunt u terugkeren naar de app: https://jureren-carnaval-sas-v3-5hv5dkb6jabwo595qmmb6e.streamlit.app/"
        "\n\nGroeten,\nKasper Tak \n\nTelefoonnummer: 06 29927267")
    msg.add_attachment(excel_bytes_1.getvalue(),
                       maintype="application",
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       filename=filename_1)
    
    msg.add_attachment(excel_bytes_2.getvalue(),
                       maintype="application",
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       filename=filename_2)
    
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(
            st.secrets["email"]["from"],
            st.secrets["email"]["app_password"])
        smtp.send_message(msg, to_addrs=[msg["To"], msg["Cc"]])
#%% df to excel defs
def df_to_excel_generic(df, sheet_name):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)
    return output


def df_to_excel_rapport(df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Uitslag"
    
    #Stijlen
    header_fill = PatternFill("solid", fgColor="D6D1B0")
    row_fill_1 = PatternFill('solid', fgColor= "C7D9ED")
    row_fill_2 = PatternFill('solid', fgColor="E6EEF7")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    
    border = Border(
        left = Side(style="thin"),
        right = Side(style="thin"),
        top = Side(style="thin"),
        bottom = Side(style="thin"))
    row = 1
    
    for categorie, groep in df.groupby("Categorie", sort=False):
        ws.merge_cells(start_row=row, start_column=1, end_row = row, end_column=10)
        cell = ws.cell(row=row, column=1, value = categorie)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        
        for col in range(1,11):
            ws.cell(row=row, column=col).border = border
        row += 1 
        
        #subheaders
        headers = ["Plaats", "Nr.", "Vereniging", "Titel", "Idee", "Bouwtechnisch", "Afwerking", "Carnavalesk", "Actie", "Totaal punten" ]
    
        for col, text in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=text)
            c.font = bold
            c.alignment = center
            c.fill = header_fill
            c.border = border
        row += 1
        
        # Data
        fill_toggle = True
        for _, r in groep.iterrows():
            fill = row_fill_1 if fill_toggle else row_fill_2
            fill_toggle = not fill_toggle
            values = [
                r["Plaats"],
                r["Nr."],
                r["Vereniging"],
                r["Titel"],
                r["Idee"],
                r["Bouwtechnisch"],
                r["Afwerking"],
                r["Carnavalesk"],
                r["Actie"],
                r["Totaal punten"]
                ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = fill
                cell.border = border
                cell.alignment = center if col != 4 else left
                # waarden totale punten dikgedrukt
                if col == 10:
                    cell.font = Font(bold=True)
            row += 1
        row += 1
        
    # kolombreedtes
    widths =[8, 6, 25, 90, 8, 12, 10, 12, 8, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w
    
    wb.save(output)
    output.seek(0)                 
    return output
    

def df_to_excel_colored(df):
    # Exporteer DataFrame tijdelijk naar BytesIO
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    
    # Stijlen
    header_fill = PatternFill("solid", fgColor="D6D1B0")
    subheader_fill = PatternFill('solid', fgColor= "E7E5D3")
    row_fill = PatternFill('solid', fgColor="E9F1FA")
    
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left = Side(style="thin"),
        right = Side(style="thin"),
        top = Side(style="thin"),
        bottom = Side(style="thin"))
    row = 1
    
    for categorie, groep in df.groupby("Categorie", sort = False):
    # Categorie-kop
        ws.cell(row=row, column = 1, value="Categorie").font = bold
        ws.cell(row=row, column = 1).fill = header_fill
        ws.cell(row=row, column = 1).border = border
        ws.merge_cells(start_row=row, start_column=2, end_row = row, end_column=3)
        cat_cell = ws.cell(row=row, column=2, value = categorie)
        cat_cell.font = bold
        cat_cell.alignment = center
        cat_cell.fill = header_fill
        cat_cell.border = border
        ws.cell(row=row, column=3).border = border
        
        row += 1
        #subkop
        headers = ["Plaats", "Vereniging", "Titel"]
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = bold
            cell.fill = subheader_fill
            cell.border = border
            cell.alignment = center if col == 1 else left
        row += 1
        #data
        for _, r in groep.iterrows():
            ws.cell(row=row, column=1, value=r["Plaats"]).alignment = center
            ws.cell(row=row, column=2, value=r["Vereniging"])
            ws.cell(row=row, column=3, value=r["Titel"])
            
            for col in range(1,4):
                ws.cell(row=row, column=col).fill = row_fill
                ws.cell(row=row, column=col).border = border
            row += 1 
        # lege rij tussen categorien
        row += 1
        
    # kolombreedtes
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 90
    
    wb.save(output)
    output.seek(0)
    
    return output
 #%% --- spul voor top-3 bepaling
# # data voor top 3
# df_existing_top3 = load_sheet_data("LeutigsteDeelnemer_2026")
# sheet_top3 = client.open("Jury_beoordelingen_2026_v1").worksheet("LeutigsteDeelnemer_2026")
# # data voor uitslag
# sheet = client.open("Jury_beoordelingen_2026_v1").worksheet("Beoordelingen_2026")

def split_categorie_nummer_titel_vereniging(combinatie):
    try: 
        categorie_part, rest = combinatie.split(" | ", 1)
        nummer_part, rest = rest.split(" - ", 1)
        vereniging_part, titel_part = rest.split(" ~ ", 1)
        
        categorie = categorie_part.strip()
        nummer = nummer_part.strip()
        vereniging = vereniging_part.strip() if vereniging_part else "Onbekend"        
        titel = titel_part.replace("**", "").strip() if titel_part else "Zonder titel"
        return categorie, nummer, titel, vereniging
    except ValueError:
        return "Onbekend", None, "Zonder titel", "Onbekend"
#%%
def login():
    st.markdown("""
    <style>
    .stApp {
        background-image: url("https://betekoppen.com/wp-content/uploads/go-x/u/90b94243-9703-4587-a11c-86168a83375e/image.png");
        background-size: 35%;
        background-repeat: no-repeat;
        background-position: center center;
        # height: 100vh;
    }
    .stApp::before {
        content: "";
        position: fixed;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        background-color: rgba(255,255,255,0.6);  /* transparante overlay */
        z-index: 0;
    }
    section.main > div {
        position: relative;
        z-index: 1;
    }
    </style>
    """, unsafe_allow_html=True)
  
    st.title("Jury login")
    
    username = st.text_input("Gebruikersnaam")
    password = st.text_input("Wachtwoord", type='password')
    if st.button("Inloggen"):
        USERS = st.secrets["users"]
        if username in USERS and USERS[username] == password:
            st.session_state["logged_in"] = True
            st.session_state['username'] = username
            st.session_state["soort"] = "w" if "_w" in username else ("g" if "_g" in username else "admin")
            st.success(f"Welkom {username}!")
            st.rerun()
        else:
            st.error("Ongeldig gebruikersnaam of wachtwoord")
            st.write('*Bij hulp: app admin via 06 29927267*')
#%%
def beoordeling_categorie_jurylid(categorie, jurylid, sheet_name="Beoordelingen_2026"):
    st.header(f'Beoordeel de stoetlopers van de categorie {categorie}')
    
    is_groepenjury = st.session_state['soort'] == 'g'
    is_wagens_categorie = "WAGEN" in categorie.upper()
    is_groepenjury_op_wagen = is_groepenjury and is_wagens_categorie
    
    if is_groepenjury_op_wagen:
        st.info("Als **groepenjury** beoordeel je bij **wagens** uitsluitend het criterium: **Carnavalesk**.")
    
    # criteria = ['Idee', 'Bouwtechnisch', 'Afwerking', 'Carnavalesk', 'Actie']
    
    # Bestaande beoordelingen ophalen 
    # sheet = client.open("Jury_beoordelingen_2026_v1").worksheet(sheet_name)
    # df_existing_beoordeling = load_sheet_data(sheet_name)
    
    sheet = st.session_state.sheet_beoordelingen
    df_existing_beoordeling = st.session_state.df_beoordelingen_cache
  
    # Filter programma op categorie
    df_tab = programma_df[programma_df['categorie'].str.contains(categorie, case=False, na=False)]
    
    # Buffer om alle wijzigingen tegelijk op te slaan
    # pending_saves = []
    
    for i, row in df_tab.iterrows():
        nummer = str(row['Nr.']).strip() if pd.notna(row['Nr.']) else "Nummer onbekend"
        vereniging = str(row['vereniging']).strip() if pd.notna(row['vereniging']) else "Onbekend"
        titel = str(row['titel']).strip() if pd.notna(row['titel']) else "Zonder titel"
        st.divider()
        st.markdown(f"#### {nummer}🎭 **{vereniging}** — *{titel}* ")

        # --- Logica voor welke criteria getoond worden ---
        if is_groepenjury_op_wagen:
            criteria = ["Carnavalesk"]
        else:
            criteria = ["Idee", "Bouwtechnisch", "Afwerking", "Carnavalesk", "Actie"]
        
        # ## als categorie Groepen (A/B/C) is en soort_jury is jury_g: dan ook Carnavalesk
        # if "G" in categorie and st.session_state['soort'] == 'g':
        #     criteria = ["Idee", "Bouwtechnisch", "Afwerking", "Carnavalesk", "Actie"]
        # ## als categorie Wagens (A/B) en soort_jury is jury_g: dan ALLEEN Carnavalesk
        # elif "W" in categorie and st.session_state['soort'] == 'g':
        #     criteria = ["Carnavalesk"]
        # ## als categorie Wagens en soort_jury is jury_w: dan ALLES criteria
        # elif "W" in categorie: #and st.session_state['soort'] == 'w':
        #     criteria = ["Idee", "Bouwtechnisch", "Afwerking", "Carnavalesk", "Actie"]
        # ## anders, (dus bij T&K en E&D), ook Carnavalesk
        # else:
        #     criteria = ["Idee", "Bouwtechnisch", "Afwerking", "Carnavalesk", "Actie"]
        
        # Controleren op dit jurylid deze deelnemer al heeft beoordeeld
        mask = (
            (df_existing_beoordeling["Jurylid"] == jurylid)
            & (df_existing_beoordeling["Deelnemer_nummer"] == nummer)
            & (df_existing_beoordeling["Deelnemer_vereniging"] == vereniging)
            & (df_existing_beoordeling["Deelnemer_titel"] == titel)
            )

        if not df_existing_beoordeling.empty and mask.any():
            bestaande_rij = df_existing_beoordeling.loc[mask].iloc[0]
            default_scores = [bestaande_rij.get(c, 5) for c in criteria]
            st.info("🟡 Bestaande beoordeling gevonden - je kunt aanpassen of bijwerken")
        else:
            default_scores = [5] * len(criteria)
            
            
        # Data tonen met bestaande of standaardwaardes via data_editor
        editor_df = pd.DataFrame({
            "Criterium": criteria,
            "Beoordeling (1-10)": default_scores})
        editor_key = f"data_editor_{i}_{jurylid}"
        editor_df = st.data_editor(
            editor_df,
            key = editor_key,
            num_rows="fixed",
            column_config={
                "Beoordeling (1-10)": st.column_config.NumberColumn(
                            "Beoordeling (1-10)",
                            min_value = 1,
                            max_value = 10,
                            step = 1)
                }
            )
        
        # Voeg opslaan toe aan buffer i.p.v. direct over te schrijven naar Google Sheets
        btn_key = f"btn_save_{i}_{jurylid}"
        if st.button(f"💾 Voeg beoordeling toe aan wachtrij ({nummer})", key=btn_key):
                def get_score(c):
                    if c in editor_df["Criterium"].values:
                        return int(editor_df.loc[editor_df["Criterium"] == c, "Beoordeling (1-10)"].values[0])
                    return 0
                
                # Dictionary aanmaken van alle scores
                new_row = {
                    "Jurylid": jurylid,
                    "Categorie": categorie,
                    "Deelnemer_nummer": nummer,
                    "Deelnemer_vereniging": vereniging,
                    "Deelnemer_titel": titel,
                    "Idee": 0 if is_groepenjury_op_wagen else get_score("Idee"),
                    "Bouwtechnisch":0 if is_groepenjury_op_wagen else get_score("Bouwtechnisch"),
                    "Afwerking": 0 if is_groepenjury_op_wagen else get_score("Afwerking"),
                    "Carnavalesk": get_score("Carnavalesk"),
                    "Actie": 0 if is_groepenjury_op_wagen else get_score("Actie"),
                    "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
                
                st.session_state['pending_saves'].append(new_row)
                st.warning(f'✅ Nieuwe beoordeling toegevoegd aan wachtrij ({nummer}). ! Let op: nog net definitief opgeslagen, dat moet onderin !')
    if st.session_state['pending_saves']:
        st.info(f"📦 {len(st.session_state['pending_saves'])} beoordelingen klaar om te uploaden")
        if st.button('📤 Alles opslaan naar Google Sheet'):
            try:
                # sheet - client.open("Jury_beoordelingen_2026_v1").worksheet(sheet_name)
                # df_existing_beoordeling = load_sheet_data(sheet_name)
                df_existing_beoordeling = st.session_state.df_beoordelingen_cache
                sheet = st.session_state.sheet_beoordelingen
                to_append = [] # nieuwe rijen verzamelen
                updates = [] # bestaande rijen (row_index, values)
                for row in st.session_state['pending_saves']:
                    mask = (
                        (df_existing_beoordeling["Jurylid"] == row['Jurylid'])
                        & (df_existing_beoordeling["Deelnemer_nummer"] == row["Deelnemer_nummer"])
                        & (df_existing_beoordeling["Deelnemer_vereniging"] == row["Deelnemer_vereniging"])
                        & (df_existing_beoordeling["Deelnemer_titel"] == row["Deelnemer_titel"])
                        )
                
                    # checken of er moet worden overgeschreven of toevoegen
                    if not df_existing_beoordeling.empty and mask.any():
                        row_index = mask[mask].index[0] + 2
                        updates.append((row_index, list(row.values())))
                        # updates.update(f"A{row_index}:J{row_index}", [list(new_row.values())])
                    else:
                        to_append.append(list(row.values()))
                        # st.success(f"✅ Nieuwe beoordeling opgeslagen voor {titel}")
                if to_append:
                    try:
                        sheet.append_rows(to_append, value_input_option="USER_ENTERED")
                    except Exception:
                        for r in to_append:
                            sheet.append_row(r, value_input_option="USER_ENTERED")
                for row_index, values in updates:
                    sheet.update(f"A{row_index}:J{row_index}", [values], value_input_option="USER_ENTERED")
                st.success(f"{len(st.session_state['pending_saves'])} beoordeling(en) succesvol opgeslagen!")
                
                if st.session_state.pending_saves:
                    new_df = pd.DataFrame(
                        st.session_state.pending_saves,
                        columns=st.session_state.df_beoordelingen_cache.columns)
                    
                    st.session_state.df_beoordelingen_cache = pd.concat(
                        [st.session_state.df_beoordelingen_cache, new_df],
                        ignore_index=True)
                
                st.session_state['pending_saves'].clear() # leegmaken na opslaan
                # load_sheet_data.clear() # cache verversen/vernieuwen
            except Exception as e:
                    st.error(f"❌ Fout bij opslaan: {e}")
    # ------- 
#%% 
if not st.session_state['logged_in']:
    login()

else:
    jurylid = st.session_state["username"]
    st.sidebar.success(f"Ingelogd als: {st.session_state['username']}")
    if st.sidebar.button("Uitloggen"):
        st.session_state["logged_in"] = False
        st.session_state["username"] = None
        st.session_state["soort"] = None
        st.rerun()
    
    # Toon het aantal nog niet opgeslagen beoordelingen
    # if st.session_state.get('pending_saves'):
    #     st.sidebar.info(f"📦 {len(st.session_state['pending_saves'])} beoordelingen in wachtrij")
    # else:
    #     st.sidebar.info("📭 Geen openstaande beoordelingen")
    @st.cache_data(ttl=300)
    def load_programma():
        df = pd.read_excel("Programma stoetopstellers 2026.xlsx", nrows=61)
        return df
    programma_df = load_programma()
    programma_df = programma_df.iloc[:, :11] # tot kolom K (11e)
    programma_df  = programma_df.rename(columns={'nr.': 'Nr.'})
    programma_df = programma_df.dropna(how='all') # alle lege rijen verwijderen
    programma_df = programma_df[programma_df['categorie'].notna() & programma_df['aantal deelnemers'].notna()].reset_index(drop=True) # alleen data met iets in Categorie en iets in Aantal Deelnemers

# logica voor tabbladen GENERIEK --------------------------------------------------------------------------------------------------------
    unique_categories = programma_df['categorie'].unique().tolist()
    tab_home = ['Home']
    tab_categories = []
    tab_last = ['Leutigste Deelnemer', 'Uitslag']
    if st.session_state['soort'] == 'admin':
        tab_last.append("Secretariaat")
    
    if st.session_state['soort'] == 'w': 
        # wagenjury: alles behalve groepen
        tab_categories += [x for x in unique_categories if "groepen" not in x]
                            
    elif st.session_state['soort'] == 'g':  
        # groepenjury: groepen + wagens
        tab_categories += [x for x in unique_categories if "groepen" in x or "wagens" in x]   
    else:
        # admin
        tab_categories = unique_categories
    tabs_total = tab_home + tab_categories + tab_last
  
    # voor tabblad met top-3
    alle_titels_met_vereniging = [
        f"{str(row['categorie']).strip()} | "
        f"{str(row['Nr.']).strip()} - "
        f"{str(row['vereniging']).strip() if pd.notna(row['vereniging']) and str(row['vereniging']).strip() != '' else 'Onbekend'} ~ "
        f"{str(row['titel']).strip() if pd.notna(row['titel']) and str(row['titel']).strip() != '' else 'Zonder titel'}"
        for _, row in programma_df.iterrows()
                            ]  
    st.title("Jury carnavalsoptocht Sas van Gent")

    if st.session_state['active_tab'] not in tabs_total:
        st.session_state['active_tab'] = 'Home'
      
    active_tab = st.session_state['active_tab']
    selected_tab = st.radio(
        "Kies tabblad:",
        tabs_total,
        index= tabs_total.index(active_tab),
        horizontal = True,
        key='radio_tabs')
    if selected_tab != active_tab:
        st.session_state['active_tab'] = selected_tab
        st.rerun()
    
    ## ---- INHOUD PER TABBLAD -------
    # Home tabblad
    if st.session_state['active_tab'] == "Home":
        st.markdown("### 🏠 Homepagina ")
        st.markdown("##### Welkom jurylid")
        with st.expander("Wat wordt er verwacht?"):
            st.write('''
                Nadat u heeft ingelogd, verschijnen de desbetreffende tabbladen voor u. Deze tabbladen bevatten de (sub)categorieën waarover u beoordelingen moet uitbrengen.
                Elk tabblad dinet u af te gaan om beoordelingen te geven over de deelnemers. Bij sommige tabbladen is het mogelijk dat deze geen deelnemers bevatten. Hier vindt dus 
                vanzelfsprekend geen beoordeling plaats. 
            ''')
        with st.expander("Hoe werkt het beoordelen?"):
            st.write('''
                De beoordelingen worden per jurylid uitgevoerd. Hierbij wordt per (sub)categorie per deelnemer beoordeeld.
                Er worden op vijf criteria beoordeeld: Idee, Bouwtechnisch, Afwerking, Carnavalesk, Actie. 
                Per (sub)categorie per deelnemer komt een tabel tevoorschijn, waarin al de standaardwaarden: 5 staan. Door op het getal te klikken, kunt u het aanpassen 
                naar uw eigen beoordeling. Wanneer u een deelnemer (dus een gehele tabel) heeft beoordeeld, dient u op de Opslaan-knop te drukken.
                Dit slaat de beoordelingen op per deelnemer/tabel.
                
            ''')
        with st.expander("Mogelijke meldingen:"):
            st.write('Er zijn verschillende meldingen die tevoorschijn kunnen komen tijdens het (opslaan van de) beoordelingen.')
            st.info("Mededeling; Niks aan de hand. Er zijn eerdere data gevonden.")
            st.warning("Waarschuwing; Er moet nog iets anders verricht worden.")
            st.error("Fout! Er is iets fout gegaan. Hier kunt u niks aan doen. Het is een fout dat de beheerder moet oplossen.")
            
        with st.expander("Hoe werkt het opslaan?"):
            st.write('''
                Wanneer de beoordelingen per deelnemer zijn ingevuld, worden deze gelijk opgeslagen in een verborgen bestand in de Cloud. 
                Naast de vijf beoordelingscriteria worden ook de gebruikersnaam, categorie, vereniging, titel en tijd opgeslagen in het bestand. 
            ''')
        with st.expander("Wanneer en hoe de uitslag?"):
            st.write('''
                De uitslag kan pas worden bepaald wanneer ieder van de juryleden op de laatste knop: 'Definitief opslaan' heeft gedrukt.
                De uitslag wordt vervolgens per categorie (Wagens, Groepen, T&K en E&D) bepaald o.b.v. de ingevulde scores.
                Ook worden de resultaten van jullie top-3 deelnemers genomen om de Leutigste Deelnemer te bepalen. 
                Uiteindelijk kan er op de knop: 'Als PDF opslaan' worden gedrukt om het overzicht van de uitslag in te zien.
            ''')
        with st.expander("De Leutigste Deelnemer"):
            st.write('''
                In het laatste tabblad wordt naar uw top-3 deelnemers gevraagd. Hierbij wordt de eerste keuze als beste beschouwd en krijgt 3 punten. De nummer 3 ontvangt 1 punt.
                Uiteindelijk worden door alle top-3-lijsten van de juryleden bepaald welke deelnemer de Leutigste Deelnemer was dit jaar!
            ''')
    
        
    for categorie in tab_categories:
        if st.session_state['active_tab'] == categorie:
            beoordeling_categorie_jurylid(categorie, jurylid)
    

    if st.session_state['active_tab'] == "Leutigste Deelnemer":
        st.header("Leutigste Deelnemer")
        st.write("Vul hier jouw top-3 deelnemers in voor *De leutigste deelnemer*. Let op: jouw ingevulde top-3 is definitief en kan niet zomaar aangepast worden!")
        # Filter bestaande keuzes van dit jurylid
        # bestaande_keuzes = df_existing_top3[df_existing_top3["Jurylid"] == jurylid]
        sheet_top3 = st.session_state.sheet_top3
        df_existing_top3 = st.session_state.df_top3_cache
        bestaande_keuzes = df_existing_top3[df_existing_top3["Jurylid"] == jurylid]
        
        if not bestaande_keuzes.empty:
            st.info("🟡 Eerdere top-3 gevonden:")
            bestaande_keuzes = bestaande_keuzes.sort_values("Punten", ascending=False)
            bestaande_keuzes_voor_display = bestaande_keuzes[["Nr.","Vereniging", "Titel", "Punten"]]
            bestaande_keuzes_voor_display.columns = ["Nr." ,"Vereniging", "Titel", "Punten"]
            bestaande_keuzes_voor_display.index = range(1, len(bestaande_keuzes_voor_display) + 1)
            bestaande_keuzes_voor_display.index.name = "Ranking"
            st.table(bestaande_keuzes_voor_display)
            st.caption("Wil je je top-3 aanpassen? Vraag de admin om reset.")
        else:
            st.info("🆕 Nog geen top-3 opgeslagen — maak nu je keuze!")

            keuze_1 = st.selectbox("Kies uw nummer 1", alle_titels_met_vereniging, key='top1_selectbox')
            keuze_2 = st.selectbox("Kies uw nummer 2", alle_titels_met_vereniging, key='top2_selectbox')
            keuze_3 = st.selectbox("Kies uw nummer 3", alle_titels_met_vereniging, key='top3_selectbox')
            
            st.divider()
            st.write("### Jouw top 3:")
            st.write(f"🥇 Nummer 1: \t{keuze_1}")
            st.write(f"🥈 Nummer 2: \t{keuze_2}")
            st.write(f"🥉 Nummer 3: \t{keuze_3}")
          
            top_keuzes = [keuze_1, keuze_2, keuze_3] 
            punten = [3, 2, 1]
            data_new_row_top3 = []

            if len(set(top_keuzes)) < 3:
                st.warning("⚠️ Elke positie in de top 3 moet een unieke titel zijn! Pas je keuzes aan voordat je opslaat.")
            else:
                # Opslaan top-3
                btn_key_top3 = f"btn_save_top3_{jurylid}"
                if st.button(f"💾 Opslaan top-3 ({jurylid})", key=btn_key_top3):
                    for keuze, pnt in zip(top_keuzes, punten):
                        categorie, nummer, titel, vereniging = split_categorie_nummer_titel_vereniging(keuze)
        
                        mask = (df_existing_top3['Jurylid'] == jurylid) & (df_existing_top3['Nr.'] == nummer)
        
                        new_row = {
                          "Jurylid": jurylid,
                          "Categorie": categorie,
                          "Nr.": nummer,
                          "Titel": titel,
                          "Vereniging": vereniging,
                          "Punten": pnt,
                          "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
        
                        if not df_existing_top3.empty and mask.any():
                            row_index = mask[mask].index[0] + 2
                            sheet_top3.update(f"A{row_index}:F{row_index}", [list(new_row.values())], value_input_option="USER_ENTERED")
                            st.info(f"🔁 Top-3 bijgewerkt voor {jurylid}")
                        else:
                            sheet_top3.append_row(list(new_row.values()), value_input_option="USER_ENTERED")
                    st.success(f"✅ Top 3 toegevoegd voor {jurylid}")          
    
    
    if st.session_state['active_tab'] == "Uitslag":
        st.header("🧮 Uitslag berekenen")
        
        sheet = st.session_state.sheet_beoordelingen
        df_existing_beoordeling = st.session_state.df_beoordelingen_cache.copy()
        
        sheet_top3 = st.session_state.sheet_top3
        df_top_3 = st.session_state.df_top3_cache.copy()
        # sheet = client.open("Jury_beoordelingen_2026_v1").worksheet("Beoordelingen_2026")
        # records = sheet.get_all_records() # normale beoordelingen
        # records_top_3 = sheet_top3.get_all_records() #leutigste deelnemer
        # df_top_3 = pd.DataFrame(records_top_3) # leutigste deelnemer
        
        df_beoordelingen = df_existing_beoordeling.copy() # normale beoordelingen
        
        alle_juryleden = [j for j in st.secrets["users"].keys() if not j.startswith("admin")]
        
        # Bepaal wie beoordelingen heeft ingeleverd
        ingeleverde_juryleden = df_beoordelingen["Jurylid"].unique().tolist()
        missende_juryleden = [j for j in alle_juryleden if j not in ingeleverde_juryleden]
        
        # Toon status
        st.subheader("Status juryleden")
        if missende_juryleden:
            st.warning("⚠️ Niet alle juryleden hebben hun beoordelingen afgerond.")
            st.write("Nog ontbrekend:", ", ".join(missende_juryleden))
        else:
            st.success("✅ Alle juryleden hebben hun beoordelingen ingeleverd!")
            
        # Forceerbare berekening
        forceer = st.checkbox("💥 Forceren (ook als nog niet alles is ingeleverd)")
        
        if (not missende_juryleden) or st.session_state['soort'] not in ['w', 'g'] or forceer:
            if st.button("📊 Bereken uitslag"):
                # st.info("Uitslag wordt berekend...")
                
                kolommen_criteria = ["Idee", "Bouwtechnisch", "Afwerking", "Carnavalesk", "Actie"]
                df = df_beoordelingen.iloc[1:]
                df["Categorie"] = df["Categorie"].str.upper()
                df_leutigste = df_top_3.iloc[1:]
                df_leutigste["Categorie"] = df_leutigste["Categorie"].str.upper()
                df = df.rename(columns={
                                "Deelnemer_nummer":'Nr.',
                                "Deelnemer_titel": "Titel",
                                "Deelnemer_vereniging": "Vereniging"})

                df_resultaten = df.copy()
                # Eerst dataframe met alleen de uitslag per categorie. Geen uitslag voor carnavalesk
                df_uitslag_categorien = df_resultaten.groupby(['Categorie', 'Nr.', 'Vereniging', 'Titel'])[kolommen_criteria].sum().reset_index()
                df_uitslag_categorien["Totaal punten"] = df_uitslag_categorien[kolommen_criteria].sum(axis=1)
                # sorteren & plaats toevoegen
                df_uitslag_categorien = df_uitslag_categorien.sort_values(by=['Categorie', 'Totaal punten'], ascending = [True, False]).reset_index(drop=True)
                df_uitslag_categorien['Plaats'] = df_uitslag_categorien.groupby("Categorie").cumcount() + 1
                df_uitslag_categorien['Beoordelingscriterium'] = 'Algemeen'

                kolomvolgorde = ["Plaats","Categorie", "Beoordelingscriterium", "Nr.", "Vereniging", "Titel", "Idee", "Bouwtechnisch", "Afwerking", "Carnavalesk", "Actie", "Totaal punten"]
                df_uitslag_categorien = df_uitslag_categorien[kolomvolgorde]

                categorie_volgorde = ["WAGENS A", "WAGENS B", "WAGENS C", 
                                      "GROEPEN A", "GROEPEN B" ,"GROEPEN C", "GROEPEN D",
                                      "TK-A", "TK-B", "TK-C",
                                      "ED-A", "ED-B", "ED-C"]
                df_uitslag_categorien["Categorie"] = pd.Categorical(df_uitslag_categorien['Categorie'], categories= categorie_volgorde,ordered=True)
                df_uitslag_categorien = df_uitslag_categorien.sort_values(by=['Categorie', 'Plaats']).reset_index(drop=True)
                df_rapport_categorien = df_uitslag_categorien.copy()
                # Recap van beoordeling Carnavaleske wagen
                df_uitslag_carnavalesk = df_rapport_categorien.sort_values(by='Carnavalesk', ascending=False).reset_index(drop=True)
                df_uitslag_carnavalesk['Plaats'] = range(1, len(df_uitslag_carnavalesk) + 1)
                df_uitslag_carnavalesk['Beoordelingscriterium'] = 'Carnavalesk'
                df_uitslag_carnavalesk['Categorie'] = 'Carnavalesk'

                df_rapport_carnavalesk = df_uitslag_carnavalesk.copy()

                # Recap van Leutigste deelnemer
                df_uitslag_leutigste = df_leutigste.groupby(["Nr.", "Vereniging", "Titel"])["Punten"].sum().reset_index()
                df_uitslag_leutigste = df_uitslag_leutigste.sort_values(by="Punten", ascending=False).reset_index(drop=True)
                df_uitslag_leutigste["Plaats"] = range(1, len(df_uitslag_leutigste) + 1)
                df_uitslag_leutigste["Beoordelingscriterium"] = "Leutigste Deelnemer"
                df_uitslag_leutigste["Categorie"] = "Leutigste Deelnemer"
                df_uitslag_leutigste = df_uitslag_leutigste.rename(columns={"Punten":"Totaal punten"})
                # beoordelingscriteria op 0 punten zetten voor leutigste deelnemer. Zo kunnen de dataframes bij elkaar gevoegd worden. (Zelfde kolommen)
                for col in kolommen_criteria:
                    if col not in df_uitslag_leutigste.columns:
                        df_uitslag_leutigste[col] = 0
                
                df_uitslag_leutigste = df_uitslag_leutigste[kolomvolgorde[:]]
                df_rapport_leutigste = df_uitslag_leutigste.copy()
                
                df_rapport = pd.concat([df_rapport_categorien, df_rapport_carnavalesk, df_rapport_leutigste], ignore_index=True)
                
                ## Nr. toevoegen vanuit Programma Stoetopstellers 
                # df_nummers = programma_df.copy()
                # df_nummers = df_nummers.rename(columns={"titel": "Titel", 
                #                                         "vereniging": "Vereniging"})
                # df_rapport_met_nummers = df_rapport.merge(df_nummers[["nr.", "Titel"]], on=["Titel"], how='left')
                # df_rapport_met_nummers = df_rapport_met_nummers.rename(columns={"nr.": "Nr."})
                kolommen_rapport_volgorde = ["Plaats", "Nr.", "Categorie", "Vereniging", "Titel", "Idee", "Bouwtechnisch", "Afwerking", "Carnavalesk", "Actie", "Totaal punten" ]
                
                # met df_rapport ipv df_rapport_met_nummers, want nummers zitten er anno 10-02-26 al in.
                df_rapport_met_nummers = df_rapport[kolommen_rapport_volgorde]
                
                # Uitslag/Rappot voor pers
                top_3_algemeen = (df_rapport_categorien.groupby("Categorie", group_keys=False).head(3))
                top_3_algemeen = top_3_algemeen[['Plaats', 'Categorie', 'Vereniging', 'Titel']]
                top_3_carnavalesk = (df_rapport_carnavalesk.head(3))
                top_3_carnavalesk['Categorie'] = 'Carnavalesk'
                top_3_carnavalesk = top_3_carnavalesk[['Plaats', 'Categorie', 'Vereniging', 'Titel']]
                top_1_leutigste = (df_rapport_leutigste.head(1))
                top_1_leutigste = top_1_leutigste[["Plaats", "Categorie", "Vereniging", "Titel"]] 

                uitslag_top_3_pers = pd.concat([top_3_algemeen, top_3_carnavalesk, top_1_leutigste], ignore_index=True)
                
                
                st.session_state.df_rapport = df_rapport_met_nummers
                st.session_state.df_pers = uitslag_top_3_pers
                st.session_state.uitslag_berekend = True
                
                st.session_state.Rapport_excel = None
                st.session_state.Pers_excel = None
                st.session_state.mail_verzonden = False
                
                # hier de berekeningslogica
            if st.session_state.uitslag_berekend:
                st.success("Uitslag is berekend")
                
                if st.session_state.Rapport_excel is None:
                    st.session_state.Rapport_excel = df_to_excel_rapport(st.session_state.df_rapport)
                st.download_button("Download rapport naar Excel", data = st.session_state.Rapport_excel, file_name="uitslag_rapport.xlsx", 
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            
                if st.session_state.Pers_excel is None:
                    st.session_state.Pers_excel = df_to_excel_colored(st.session_state.df_pers)
                st.download_button("Download persuitslag", data=st.session_state.Pers_excel, file_name="pers_uitslag.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    
                    
                # excel_buffer = df_to_excel_colored(df_rapport)
                # st.download_button(
                #     label = "Download rapport als Excel-bestand",
                #     data = excel_buffer,
                #     file_name= "uitslag_rapport.xlsx",
                #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                # excel_buffer_2 = df_to_excel_colored(uitslag_top_3_pers)
                # st.download_button(
                #     label = "Pers-uitslag",
                #     data = excel_buffer_2,
                #     file_name= "pers_uitslag.xlsx",
                #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                
            if st.session_state.uitslag_berekend:
                if st.button("Verstuur rapport via mail"):
                    mail_excel(st.session_state.Rapport_excel, "Volledig_rapport_uitslag.xlsx",
                               st.session_state.Pers_excel, "Persuitslag.xlsx")
                    st.session_state.mail_verzonden = True
                    st.success("Mail succesvol verzonden!")
                

      
                
        else:
            st.info("⏳ Wacht op alle juryleden, of vink 'forceren' aan om toch te berekenen.")
            
            
    if st.session_state['active_tab'] == "Secretariaat":
        st.header("Secretariaat: volledige beoordelingen")
        st.write("Hier kan de volledige beoordeling worden gedownload en vervolgens worden geüpload om nieuwe uitslag te creëeren.")
        
        # ------------------------ Normale beoordelingen -------------------------------
        st.subheader("Jurybeoordelingen (punten per criterium)")
        
        excel_buffer_beoordelingen = df_to_excel_generic(st.session_state.df_beoordelingen_cache, sheet_name="Beoordelingen")
        
        st.download_button(
            label = "Download volledige beoordelingen",
            data = excel_buffer_beoordelingen,
            file_name = "Beoordelingen_volledig.xlsx",
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        uploaded_beoordelingen = st.file_uploader("Upload bijgewerkte beoordelingen", type = ["xlsx"], key="upload_beoordelingen")
        
        if uploaded_beoordelingen is not None:
            try:
                df_new = pd.read_excel(uploaded_beoordelingen)
                expected_cols = st.session_state.df_beoordelingen_cache.columns.tolist()
                if list(df_new.columns) != expected_cols:
                    st.error(f"Kolommen komen niet overeen. Verwacht: {expected_cols}")
                    st.code(expected_cols)
                else:
                    st.session_state.df_beoordelingen_cache = df_new
                    st.success("Nieuwe beoordelingen succesvol ingeladen! Je kunt de uitslag berekenen.")
            except Exception as e:
                st.error(f"Fout bij upload: {e}")
        
        st.divider()
        # ------------------------ Leutigste Deelnemer -------------------------------
        st.subheader("Leutigste Deelnemer (top-3 per jurylid)")
        
        excel_buffer_top3 = df_to_excel_generic(st.session_state.df_top3_cache, sheet_name="LeutigsteDeelnemer")
        
        st.download_button(
            label = "Download Leutigste Deelnemer beoordelingen",
            data = excel_buffer_top3,
            file_name = "LeutigsteDeelnemer_volledig.xlsx",
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        uploaded_top3= st.file_uploader("Upload bijgewerkte beoordelingen", type = ["xlsx"], key="upload_top3")
        
        if uploaded_top3 is not None:
            try:
                df_top3_new = pd.read_excel(uploaded_top3)
                expected_cols_top3 = st.session_state.df_top3_cache.columns.tolist()
                if list(df_top3_new.columns) != expected_cols_top3:
                    st.error(f"Kolommen komen niet overeen. Verwacht: {expected_cols_top3}")
                    st.code(expected_cols_top3)
                else:
                    st.session_state.df_top3_cache = df_top3_new 
                    st.success("Nieuwe beoordelingen succesvol ingeladen! Je kunt de uitslag berekenen.")
            except Exception as e:
                st.error(f"Fout bij upload: {e}")
                
        st.info("Let op: wijzigingen hier **overschrijven de jury-invoer**."
                "Na upload kun je direct naar het tabblad **Uitslag** om opnieuw te berekenen.")
        


